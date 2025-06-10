import streamlit as st
import sqlite3
from datetime import datetime
import pandas as pd
import os
from werkzeug.utils import secure_filename
import base64
from image_analyzer import extract_text_from_image, analyze_document_content
import logging
import traceback  # Add this import
from document_processor import UnifiedDocumentProcessor  # Add this import
import re  # Add this import
import openpyxl  # For xlsx files
from docx import Document as DocxDocument  # For docx files
from PIL import Image  # For image files

# Constants
UPLOAD_FOLDER = "uploads"
ALLOWED_EXTENSIONS = {'pdf', 'xlsx', 'docx', 'jpg', 'png'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

logger = logging.getLogger(__name__)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_db():
    return sqlite3.connect('document_management.db')

def init_db():
    conn = get_db()
    c = conn.cursor()
    
    # Create tables
    c.execute('''CREATE TABLE IF NOT EXISTS departments
                 (id INTEGER PRIMARY KEY, name TEXT UNIQUE)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS projects
                 (id INTEGER PRIMARY KEY, name TEXT UNIQUE)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS archive_codes
                 (id INTEGER PRIMARY KEY, code TEXT UNIQUE)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS documents
                 (id INTEGER PRIMARY KEY,
                  title TEXT,
                  file_title TEXT,
                  description TEXT,
                  doc_date DATE,
                  end_date DATE,
                  doc_number TEXT,
                  alt_number TEXT,
                  department_id INTEGER,
                  project_id INTEGER,
                  archive_code_id INTEGER,
                  security_class TEXT,
                  status TEXT,
                  created_at TIMESTAMP,
                  file_paths TEXT,
                  file_names TEXT,
                  FOREIGN KEY (department_id) REFERENCES departments (id),
                  FOREIGN KEY (project_id) REFERENCES projects (id),
                  FOREIGN KEY (archive_code_id) REFERENCES archive_codes (id))''')
    
    # Insert initial data
    departments = [
        "Sekretaris Perusahaan (SEP)", "Akuntansi dan Keuangan (AKK)",
        "Pengendalian dan Administrasi (PDA)", "Human Capital and General Affair (HCG)",
        "Supply Chain dan Peralatan (SCP)", "Satuan Pengawas Internal (SPI)",
        "Pemasaran (PEM)", "Sistem dan Informasi Teknologi (SIT)",
        "Manajemen Risiko (MAR)", "Perencanaan Korporasi (PEK)",
        "Teknik dan Desain (TED)", "Quality, Health, Safety, Security (QHS)",
        "Infra (INF) Proyek (PRO)"
    ]
    
    c.executemany('INSERT OR IGNORE INTO departments (name) VALUES (?)',
                  [(d,) for d in departments])
    
    projects = [
        "TIP Medan - Binjai", "Medan - Binjai", "Binjai - Pangkalan Brandan",
        "Junction - Palindra", "Indralaya - Prabumulih", "Palembang - Indralaya",
        "Pekanbaru - Bangkinang", "TIP Pekanbaru - Dumai", "Pekanbaru Ringroad",
        "Bangkinang - Pangkalan", "TIP Bengkulu", "Bengkulu - Taba Penanjung",
        "TIP Bakauheni - Terbanggi Besar", "TBPPKA Dukon Non Dukon", "13 Departemen"
    ]
    
    c.executemany('INSERT OR IGNORE INTO projects (name) VALUES (?)',
                  [(p,) for p in projects])
    
    archive_codes = [
        "PRO100", "PRO100.01", "PRO100.02", "PRO200", "PRO300", "PRO400",
        "PRO500", "PRO600"
    ]
    
    c.executemany('INSERT OR IGNORE INTO archive_codes (code) VALUES (?)',
                  [(a,) for a in archive_codes])
    
    conn.commit()
    conn.close()

def save_uploaded_files(files):
    if not files:
        return [], []
    
    saved_paths = []
    saved_names = []
    
    for file in files:
        if file and allowed_file(file.name):
            filename = secure_filename(file.name)
            file_path = os.path.join(UPLOAD_FOLDER, filename)
            
            with open(file_path, "wb") as f:
                f.write(file.getbuffer())
            
            saved_paths.append(file_path)
            saved_names.append(filename)
    
    return saved_paths, saved_names

def analyze_uploaded_document(file):
    """Analyze construction document with enhanced metadata extraction"""
    try:
        text = ""
        if file.type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":  # xlsx
            wb = openpyxl.load_workbook(file)
            
            # First try to find and process COVER sheet
            cover_sheet = None
            cover_variants = ['COVER', 'Cover', 'cover', 'SAMPUL', 'Sampul']
            
            for variant in cover_variants:
                if variant in wb.sheetnames:
                    cover_sheet = wb[variant]
                    break
            
            if cover_sheet:
                # Process COVER sheet first
                text += "=== COVER SHEET ===\n"
                # Read first 10 rows of COVER sheet (usually contains important metadata)
                for row in list(cover_sheet.rows)[:10]:
                    row_text = " ".join(str(cell.value or '').strip() for cell in row if cell.value)
                    if row_text:
                        text += row_text + "\n"
                text += "\n"
            
            # Then process other sheets
            for sheet_name in wb.sheetnames:
                if sheet_name not in cover_variants:
                    sheet = wb[sheet_name]
                    text += f"\n=== {sheet_name} ===\n"
                    for row in sheet.rows:
                        row_text = " ".join(str(cell.value or '').strip() for cell in row if cell.value)
                        if row_text:
                            text += row_text + "\n"
                    text += "\n"
        
        elif file.type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":  # docx
            doc = DocxDocument(file)
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            
        elif file.type.startswith('image/'):  # jpg, png
            text = extract_text_from_image(file)
            
        elif file.type == 'application/pdf':  # pdf
            text = analyze_document_content(file)
            
        else:
            raise ValueError(f"Unsupported file type: {file.type}")

        # Enhanced metadata extraction for construction documents
        metadata = {
            'title': '',
            'file_title': os.path.splitext(file.name)[0],
            'description': '',
            'doc_date': datetime.now(),
            'doc_number': '',
            'alt_number': '',
            'archive_code': '',
            'department': '',
            'project': '',
            'security_class': 'Biasa/Umum/Terbuka',
            'document_type': '',  # New field for construction doc type
            'revision': '',       # New field for revision number
            'phase': '',         # New field for project phase
            'discipline': ''     # New field for engineering discipline
        }

        # Extract archive code from filename
        filename = file.name.upper()
        archive_code_match = re.search(r'(PRO\d{3}(?:\.\d{2})?)', filename)
        if archive_code_match:
            metadata['archive_code'] = archive_code_match.group(1)

        # Department abbreviation mapping
        dept_abbrev = {
            'SEP': 'Sekretaris Perusahaan (SEP)',
            'AKK': 'Akuntansi dan Keuangan (AKK)',
            'PDA': 'Pengendalian dan Administrasi (PDA)',
            'HCG': 'Human Capital and General Affair (HCG)',
            'SCP': 'Supply Chain dan Peralatan (SCP)',
            'SPI': 'Satuan Pengawas Internal (SPI)',
            'PEM': 'Pemasaran (PEM)',
            'SIT': 'Sistem dan Informasi Teknologi (SIT)',
            'MAR': 'Manajemen Risiko (MAR)',
            'PEK': 'Perencanaan Korporasi (PEK)',
            'TED': 'Teknik dan Desain (TED)',
            'QHS': 'Quality, Health, Safety, Security (QHS)',
            'INF': 'Infra (INF) Proyek (PRO)'
        }

        # Try to find department from filename first
        for abbrev, full_name in dept_abbrev.items():
            if abbrev in filename:
                metadata['department'] = full_name
                break

        if text:
            lines = text.split('\n')
            lines = [line.strip() for line in lines if line.strip()]
            
            # If this is an Excel file with COVER sheet, prioritize its content
            if "=== COVER SHEET ===" in text:
                cover_section = text.split("=== COVER SHEET ===")[1].split("===")[0]
                cover_lines = [line.strip() for line in cover_section.split('\n') if line.strip()]
                
                # Usually title is in the first few non-empty lines of COVER sheet
                if cover_lines:
                    metadata['title'] = cover_lines[0]
                    # Description often follows the title
                    if len(cover_lines) > 1:
                        metadata['description'] = ' '.join(cover_lines[1:3])
            else:
                # Fallback to regular title extraction
                metadata['title'] = next((line for line in lines[:3] if line.isupper()), lines[0])
                
                # Extract document number (common formats: XXX-YYY-ZZZ, XX.YY.ZZ)
                doc_number_pattern = r'(?i)(?:no|number|doc|drawing)[\s.:]*([A-Z0-9\-\.]+)'
                if doc_number_match := re.search(doc_number_pattern, text):
                    metadata['doc_number'] = doc_number_match.group(1)

                # Extract revision number
                rev_pattern = r'(?i)(?:rev|revision)[\s.:]*([A-Z0-9]+)'
                if rev_match := re.search(rev_pattern, text):
                    metadata['revision'] = rev_match.group(1)

                # Detect document type based on content keywords
                doc_types = {
                    'drawing': ['drawing', 'gambar', 'layout', 'detail'],
                    'specification': ['specification', 'spesifikasi', 'spec', 'requirement'],
                    'report': ['report', 'laporan', 'analisis', 'analysis'],
                    'calculation': ['calculation', 'perhitungan', 'analysis'],
                    'method_statement': ['method', 'procedure', 'prosedur', 'metode']
                }
                
                for doc_type, keywords in doc_types.items():
                    if any(keyword in text.lower() for keyword in keywords):
                        metadata['document_type'] = doc_type
                        break

                # Extract project phase
                phases = ['design', 'tender', 'construction', 'as-built', 'preliminary']
                for phase in phases:
                    if phase in text.lower():
                        metadata['phase'] = phase
                        break

                # Extract description from early paragraphs
                desc_lines = []
                for line in lines[1:4]:
                    if not any(pattern in line.lower() for pattern in ['rev', 'no.', 'date']):
                        desc_lines.append(line)
                metadata['description'] = ' '.join(desc_lines)

                # Try to find department in content if not found in filename
                if not metadata['department']:
                    for abbrev, full_name in dept_abbrev.items():
                        if any(abbrev in line.upper() for line in lines[:10]):
                            metadata['department'] = full_name
                            break
                
                # Try to find archive code in content if not found in filename
                if not metadata['archive_code']:
                    pro_pattern = r'(PRO\d{3}(?:\.\d{2})?)'
                    for line in lines[:10]:
                        if code_match := re.search(pro_pattern, line.upper()):
                            metadata['archive_code'] = code_match.group(1)
                            break

        return metadata
        
    except Exception as e:
        logger.error(f"Document analysis error: {str(e)}")
        logger.error(traceback.format_exc())
        return None

def create_document_record(metadata: dict, file=None):
    """Create document record with given metadata"""
    conn = get_db()
    try:
        # Save file if provided
        saved_paths = []
        saved_names = []
        if file:
            paths, names = save_uploaded_files([file])
            saved_paths.extend(paths)
            saved_names.extend(names)
        
        # Insert record
        c = conn.cursor()
        c.execute('''
            INSERT INTO documents 
            (title, file_title, description, doc_date, 
             doc_number, alt_number, department_id, project_id,
             archive_code_id, security_class, status, created_at,
             file_paths, file_names)
            VALUES (?, ?, ?, ?, ?, ?, 
                    (SELECT id FROM departments WHERE name = ?),
                    (SELECT id FROM projects WHERE name = ?),
                    (SELECT id FROM archive_codes WHERE code = ?),
                    ?, ?, ?, ?, ?)
        ''', (
            metadata['title'],
            metadata['file_title'],
            metadata['description'],
            metadata['doc_date'],
            metadata['doc_number'],
            metadata['alt_number'],
            metadata['department'],
            metadata['project'],
            metadata['archive_code'],
            metadata['security_class'],
            metadata['status'],
            datetime.now(),
            "|".join(saved_paths) if saved_paths else "",
            "|".join(saved_names) if saved_names else ""
        ))
        
        conn.commit()
        return True
        
    except Exception as e:
        logger.error(f"Error creating document record: {str(e)}")
        conn.rollback()
        raise
    finally:
        conn.close()

def create_document():
    """Create a new document record with admin authorization and auto-filling"""
    # Check admin authentication
    if not st.session_state.get('admin_authenticated', False):
        st.error("⚠️ Only administrators can create documents")
        return
        
    conn = get_db()
    try:
        # Get lookup data
        departments = pd.read_sql('SELECT * FROM departments', conn)
        projects = pd.read_sql('SELECT * FROM projects', conn)
        archive_codes = pd.read_sql('SELECT * FROM archive_codes', conn)
        
        # File upload first for analysis
        uploaded_files = st.file_uploader(
            "Upload Documents (Required)", 
            accept_multiple_files=True,
            type=list(ALLOWED_EXTENSIONS),
            help="Upload at least one document file"
        )
        
        if uploaded_files:
            # Analyze first uploaded file for metadata with detailed analysis
            with st.spinner("📄 Analyzing document content..."):
                # Get metadata using the enhanced system prompt
                metadata = analyze_uploaded_document(uploaded_files[0])
                
                if metadata:
                    st.success("✅ Document analysis complete!")
                    
                    # Show analysis results
                    with st.expander("View Analysis Results"):
                        st.json(metadata)
            
            # Create form with pre-filled data
            with st.form("create_document_form"):
                col1, col2 = st.columns(2)
                
                with col1:
                    # ...existing column 1 inputs...
                    title = st.text_input(
                        "Title*", 
                        value=metadata.get('title', '') if metadata else '',
                        placeholder="Enter document title"
                    )
                    
                    file_title = st.text_input(
                        "File Title*", 
                        value=metadata.get('file_title', '') if metadata else '',
                        placeholder="Enter file title"
                    )
                    
                    description = st.text_area(
                        "Description", 
                        value=metadata.get('description', '') if metadata else '',
                        placeholder="Enter document description"
                    )
                    
                    doc_number = st.text_input(
                        "Document Number*", 
                        value=metadata.get('doc_number', '') if metadata else '',
                        placeholder="Enter document number"
                    )
                    
                    alt_number = st.text_input(
                        "Alternative Number", 
                        value=metadata.get('alt_number', '') if metadata else '',
                        placeholder="Enter alternative number"
                    )
                    
                    # Pre-select department if found in metadata
                    dept_index = None
                    if metadata.get('department'):
                        dept_list = departments['name'].tolist()
                        try:
                            dept_index = dept_list.index(metadata['department'])
                        except ValueError:
                            pass
                    
                    department = st.selectbox(
                        "Department*", 
                        departments['name'],
                        index=dept_index,
                        placeholder="Select department"
                    )
                    
                    project = st.selectbox(
                        "Project*", 
                        projects['name'],
                        index=None,
                        placeholder="Select project"
                    )
                
                with col2:
                    # ...existing column 2 inputs...
                    doc_date = st.date_input(
                        "Document Date*",
                        value=metadata.get('doc_date', datetime.now()) if metadata else datetime.now()
                    )
                    
                    end_date = st.date_input(
                        "End Date",
                        value=metadata.get('end_date', datetime.now()) if metadata else datetime.now()
                    )
                    
                    # Pre-select archive code if found in metadata
                    code_index = None
                    if metadata.get('archive_code'):
                        code_list = archive_codes['code'].tolist()
                        try:
                            code_index = code_list.index(metadata['archive_code'])
                        except ValueError:
                            # If exact match not found, try to find closest match
                            if metadata['archive_code'].startswith('PRO'):
                                matching_codes = [i for i, code in enumerate(code_list) 
                                               if code.startswith(metadata['archive_code'][:6])]
                                if matching_codes:
                                    code_index = matching_codes[0]
                        
                    archive_code = st.selectbox(
                        "Archive Code*",
                        archive_codes['code'],
                        index=code_index,
                        placeholder="Select archive code"
                    )
                    
                    security_class = st.selectbox(
                        "Security Classification*",
                        ["Biasa/Umum/Terbuka", "Terbatas"],
                        index=0,
                        placeholder="Select security classification"
                    )
                    
                    status = st.selectbox(
                        "Status*",
                        ["Disetujui", "Versi Akhir", "Diterbitkan untuk Konstruksi", "As Built"],
                        index=0,
                        placeholder="Select status"
                    )
                
                # Create button with validation
                if st.form_submit_button("Create Document", type="primary"):
                    if not all([title, department, project, archive_code, status, uploaded_files]):
                        st.error("Please fill in all required fields")
                        return
                        
                    try:
                        # Save files and create record
                        saved_paths, saved_names = save_uploaded_files(uploaded_files)
                        
                        c = conn.cursor()
                        c.execute('''
                            INSERT INTO documents 
                            (title, file_title, description, doc_date, end_date,
                             doc_number, alt_number, department_id, project_id,
                             archive_code_id, security_class, status, created_at,
                             file_paths, file_names)
                            VALUES (?, ?, ?, ?, ?, ?, ?, 
                                    (SELECT id FROM departments WHERE name = ?),
                                    (SELECT id FROM projects WHERE name = ?),
                                    (SELECT id FROM archive_codes WHERE code = ?),
                                    ?, ?, ?, ?, ?)
                        ''', (
                            title, file_title, description, doc_date, end_date,
                            doc_number, alt_number, department, project, archive_code,
                            security_class, status, datetime.now(),
                            "|".join(saved_paths), "|".join(saved_names)
                        ))
                        
                        conn.commit()
                        st.success("✅ Document successfully created!")
                        
                        if st.button("Create Another Document", key="create_another"):
                            st.rerun()
                            
                    except Exception as e:
                        conn.rollback()
                        st.error(f"Error creating document: {str(e)}")
                        logger.error(f"Create error: {str(e)}")
                        logger.error(traceback.format_exc())
        else:
            st.info("Please upload at least one document to proceed")
            
    except Exception as e:
        st.error(f"Error: {str(e)}")
        logger.error(traceback.format_exc())
    finally:
        conn.close()

def read_documents():
    """Display existing documents with filters and text values"""
    conn = get_db()
    try:
        # Get lookup data with joins to show text values
        query = """
            SELECT 
                d.id,
                d.title,
                d.file_title,
                d.description,
                date(d.doc_date) as doc_date,
                date(d.end_date) as end_date,
                d.doc_number,
                d.alt_number,
                d.security_class,
                d.status,
                datetime(d.created_at) as created_at,
                dept.name as department,
                p.name as project,
                ac.code as archive_code
            FROM documents d
            LEFT JOIN departments dept ON d.department_id = dept.id
            LEFT JOIN projects p ON d.project_id = p.id
            LEFT JOIN archive_codes ac ON d.archive_code_id = ac.id
            WHERE 1=1
        """
        
        # Get filter options
        departments = pd.read_sql('SELECT * FROM departments', conn)
        projects = pd.read_sql('SELECT * FROM projects', conn)
        archive_codes = pd.read_sql('SELECT * FROM archive_codes', conn)
        
        # Filters
        col1, col2 = st.columns(2)
        
        with col1:
            department_filter = st.selectbox("Filter by Department", ["All"] + departments['name'].tolist())
            project_filter = st.selectbox("Filter by Project", ["All"] + projects['name'].tolist())
        
        with col2:
            archive_code_filter = st.selectbox("Filter by Archive Code", ["All"] + archive_codes['code'].tolist())
            status_filter = st.selectbox("Filter by Status", ["All", "Disetujui", "Versi Akhir", "Diterbitkan untuk Konstruksi", "As Built"])
        
        # Apply filters to query
        params = []
        if department_filter != "All":
            query += " AND dept.name = ?"
            params.append(department_filter)
        
        if project_filter != "All":
            query += " AND p.name = ?"
            params.append(project_filter)
        
        if archive_code_filter != "All":
            query += " AND ac.code = ?"
            params.append(archive_code_filter)
        
        if status_filter != "All":
            query += " AND d.status = ?"
            params.append(status_filter)
        
        # Execute query and get results
        documents = pd.read_sql(query, conn, params=params, parse_dates=['doc_date', 'end_date', 'created_at'])
        
        if not documents.empty:
            # Add numbered index column
            documents.index = range(1, len(documents) + 1)
            documents.index.name = 'No.'
            
            # Format dates properly
            for date_col in ['doc_date', 'end_date', 'created_at']:
                if date_col in documents.columns:
                    documents[date_col] = pd.to_datetime(documents[date_col]).dt.strftime('%Y-%m-%d')
            
            # Reorder columns
            display_columns = [
                'title', 'file_title', 'description', 'doc_date', 
                'end_date', 'doc_number', 'alt_number',
                'department', 'project', 'archive_code',
                'security_class', 'status', 'created_at'
            ]
            
            # Display dataframe
            st.dataframe(
                documents[display_columns],
                use_container_width=True
            )
            
            # Export functionality
            if st.button("Export to CSV"):
                export_df = documents[display_columns].copy()
                export_df.reset_index(names=['No.'], inplace=True)
                csv = export_df.to_csv(index=False)
                st.download_button(
                    "Download CSV",
                    csv,
                    "documents_export.csv",
                    "text/csv",
                    key='download-csv'
                )
        else:
            st.info("No documents found matching the filters")
            
    except Exception as e:
        st.error(f"Error reading documents: {str(e)}")
        logger.error(traceback.format_exc())
    finally:
        conn.close()

def update_document():
    """Update an existing document"""
    conn = get_db()
    try:
        # Get lookup data with proper data types
        documents = pd.read_sql('SELECT * FROM documents', conn)
        departments = pd.read_sql('SELECT * FROM departments', conn)
        projects = pd.read_sql('SELECT * FROM projects', conn)
        archive_codes = pd.read_sql('SELECT * FROM archive_codes', conn)
        
        # Convert IDs to regular Python integers
        documents['id'] = documents['id'].astype(int)
        documents['department_id'] = documents['department_id'].astype(int)
        documents['project_id'] = documents['project_id'].astype(int)
        documents['archive_code_id'] = documents['archive_code_id'].astype(int)
        
        # Select document to update
        document_ids = documents['id'].tolist()
        selected_id = st.selectbox(
            "Select Document to Update",
            options=document_ids,
            format_func=lambda x: f"{documents[documents['id'] == x]['title'].iloc[0]} (ID: {x})"
        )
        
        if selected_id:
            document = documents[documents['id'] == selected_id].iloc[0]
            
            # Form inputs
            col1, col2 = st.columns(2)
            
            with col1:
                title = st.text_input("Title", value=document['title'])
                file_title = st.text_input("File Title", value=document['file_title'])
                doc_number = st.text_input("Document Number", value=document['doc_number'] or '')
                alt_number = st.text_input("Alternative Number", value=document['alt_number'] or '')
                
                # Get current department index
                dept_index = departments[departments['id'] == document['department_id']].index[0]
                department = st.selectbox(
                    "Department",
                    options=departments['name'].tolist(),
                    index=int(dept_index)
                )
                
                # Get current project index
                proj_index = projects[projects['id'] == document['project_id']].index[0]
                project = st.selectbox(
                    "Project",
                    options=projects['name'].tolist(),
                    index=int(proj_index)
                )
            
            with col2:
                description = st.text_area("Description", value=document['description'] or '')
                
                # Handle date fields properly
                doc_date = st.date_input(
                    "Document Date",
                    value=pd.to_datetime(document['doc_date']).date() if pd.notnull(document['doc_date']) else datetime.now().date()
                )
                end_date = st.date_input(
                    "End Date",
                    value=pd.to_datetime(document['end_date']).date() if pd.notnull(document['end_date']) else datetime.now().date()
                )
                
                # Get current archive code index
                code_index = archive_codes[archive_codes['id'] == document['archive_code_id']].index[0]
                archive_code = st.selectbox(
                    "Archive Code",
                    options=archive_codes['code'].tolist(),
                    index=int(code_index)
                )
                
                security_class = st.selectbox(
                    "Security Classification",
                    options=["Biasa/Umum/Terbuka", "Terbatas"],
                    index=0 if document['security_class'] == "Biasa/Umum/Terbuka" else 1
                )
                
                status = st.selectbox(
                    "Status",
                    options=["Disetujui", "Versi Akhir", "Diterbitkan untuk Konstruksi", "As Built"],
                    index=["Disetujui", "Versi Akhir", "Diterbitkan untuk Konstruksi", "As Built"].index(document['status'])
                )
            
            if st.button("Update Document"):
                try:
                    c = conn.cursor()
                    c.execute('''
                        UPDATE documents
                        SET title = ?,
                            file_title = ?,
                            description = ?,
                            doc_date = ?,
                            end_date = ?,
                            doc_number = ?,
                            alt_number = ?,
                            department_id = (SELECT id FROM departments WHERE name = ?),
                            project_id = (SELECT id FROM projects WHERE name = ?),
                            archive_code_id = (SELECT id FROM archive_codes WHERE code = ?),
                            security_class = ?,
                            status = ?
                        WHERE id = ?
                    ''', (
                        title,
                        file_title,
                        description,
                        doc_date.strftime('%Y-%m-%d'),
                        end_date.strftime('%Y-%m-%d'),
                        doc_number,
                        alt_number,
                        department,
                        project,
                        archive_code,
                        security_class,
                        status,
                        int(selected_id)  # Convert to regular Python int
                    ))
                    
                    conn.commit()
                    st.success("✅ Document successfully updated!")
                    
                    # Add refresh button
                    if st.button("Refresh"):
                        st.rerun()
                        
                except Exception as e:
                    conn.rollback()
                    st.error(f"Error updating document: {str(e)}")
                    logger.error(f"Update error: {str(e)}")
                    logger.error(traceback.format_exc())
                    
    except Exception as e:
        st.error(f"Error loading document data: {str(e)}")
        logger.error(traceback.format_exc())
    finally:
        conn.close()

def delete_document():
    """Delete documents with bulk delete support"""
    if not st.session_state.get('admin_authenticated', False):
        st.error("⚠️ Only administrators can delete documents")
        return
        
    conn = get_db()
    try:
        # Get all documents
        docs_df = pd.read_sql('''
            SELECT d.id, d.title, d.file_title, d.doc_date, 
                   dept.name as department, d.file_paths
            FROM documents d
            LEFT JOIN departments dept ON d.department_id = dept.id
        ''', conn)
        
        if docs_df.empty:
            st.info("No documents available to delete")
            return
        
        # Add multiselect for bulk deletion
        st.warning("⚠️ Warning: Deletion cannot be undone!")
        selected_docs = st.multiselect(
            "Select Documents to Delete",
            options=docs_df['id'].tolist(),
            format_func=lambda x: (
                f"{docs_df[docs_df['id'] == x]['title'].iloc[0]} "
                f"({docs_df[docs_df['id'] == x]['department'].iloc[0]})"
            )
        )
        
        if selected_docs:
            st.info(f"Selected {len(selected_docs)} documents for deletion")
            
            # Add confirmation requirements
            col1, col2 = st.columns(2)
            with col1:
                confirm_check = st.checkbox("I understand this action is permanent")
            with col2:
                confirm_text = st.text_input("Type 'DELETE' to confirm")
            
            if st.button("Delete Selected Documents", type="primary"):
                if not (confirm_check and confirm_text == "DELETE"):
                    st.error("Please complete all confirmation steps")
                    return
                    
                try:
                    c = conn.cursor()
                    deleted_count = 0
                    
                    # Delete files first
                    for doc_id in selected_docs:
                        doc_info = docs_df[docs_df['id'] == doc_id].iloc[0]
                        
                        # Delete associated files
                        if doc_info['file_paths']:
                            for file_path in doc_info['file_paths'].split('|'):
                                if os.path.exists(file_path):
                                    os.remove(file_path)
                                    
                        # Delete markdown file if exists
                        markdown_path = os.path.join(
                            "extracted_text", 
                            f"{doc_info['file_title']}.md"
                        )
                        if os.path.exists(markdown_path):
                            os.remove(markdown_path)
                        
                        # Delete database record
                        c.execute('DELETE FROM documents WHERE id = ?', (doc_id,))
                        deleted_count += 1
                    
                    conn.commit()
                    st.success(f"✅ Successfully deleted {deleted_count} documents!")
                    st.rerun()  # Force page refresh
                    
                except Exception as e:
                    conn.rollback()
                    st.error(f"Error during bulk deletion: {str(e)}")
                    
        else:
            st.info("Select documents to delete from the list above")
            
    except Exception as e:
        st.error(f"Error in delete operation: {str(e)}")
    finally:
        conn.close()

def create_container_with_color(id, color="#E4F2EC"):
    """Create a custom styled container"""
    plh = st.container()
    html_code = """<div id = 'my_div_outer'></div>"""
    st.markdown(html_code, unsafe_allow_html=True)
    with plh:
        inner_html_code = """<div id = 'my_div_inner_%s'></div>""" % id
        plh.markdown(inner_html_code, unsafe_allow_html=True)
    
    chat_plh_style = """
        <style>
            div[data-testid='stVerticalBlock']:has(div#my_div_inner_%s):not(:has(div#my_div_outer)) {
                background-color: %s;
                border-radius: 10px;
                padding: 20px;
                margin: 10px 0;
                box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            };
        </style>
        """
    st.markdown(chat_plh_style % (id, color), unsafe_allow_html=True)
    return plh

def show_landing_page():
    """Placeholder to prevent import errors"""
    pass

# Make functions explicitly available for import
__all__ = ['init_db', 'create_document', 'read_documents', 'update_document', 'delete_document']

# Only run main() if script is run directly
if __name__ == "__main__":
    st.title("Records Management System")
    init_db()
    
    menu = ["Create", "Read", "Update", "Delete"]
    choice = st.selectbox("Operation", menu)
    
    if choice == "Create":
        st.subheader("Add New Document")
        create_document()
    elif choice == "Read":
        st.subheader("View Documents")
        read_documents()
    elif choice == "Update":
        st.subheader("Update Document")
        update_document()
    elif choice == "Delete":
        st.subheader("Delete Document")
        delete_document()